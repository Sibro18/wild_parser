import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from typing import List
from models import ProductDetails


def save_to_excel(products: List[ProductDetails], filename: str = "report.xlsx"):
	"""Save a list of ProductDetails into a formatted Excel report."""

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Wildberries Report"

	# Column definitions
	column_idx_dict = {
		"URL": 1,
		"Артикул": 2,
		"Название": 3,
		"Описание": 4,
		"Продавец": 5,
		"URL продавца": 6,
		"Рейтинг": 7,
		"Кол-во отзывов": 8,
		"Кол-во": 9,
		"Размеры": 10,
		"Цена (по размерам)": 11,
		"Изображения": 12,
	}
	ws.append(list(column_idx_dict.keys()))

	# Header styling
	header_font = Font(bold=True)
	header_fill = PatternFill(
		start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"
	)
	thin_border = Border(
		left=Side(style="thin"),
		right=Side(style="thin"),
		top=Side(style="thin"),
		bottom=Side(style="thin"),
	)

	for idx in list(column_idx_dict.values()):
		cell = ws.cell(row=1, column=idx)
		cell.font = header_font
		cell.fill = header_fill
		cell.border = thin_border
		cell.alignment = Alignment(horizontal="center", vertical="center")

	wrap_alignment = Alignment(wrap_text=True, vertical="top")

	for row_idx, p in enumerate(products, start=2):
		price_lines = [
			f"{size}: {pr.basic}/{pr.product}" for size, pr in p.price.items()
		]
		row = [
			p.url,
			p.article_id,
			p.name,
			p.description or "",
			p.seller_name,
			p.seller_url,
			p.rating,
			p.feedbacks_count,
			p.quantity,
			", ".join(p.sizes),
			"\n".join(price_lines),
			"\n".join(p.image_urls) if p.image_urls else "",
		]
		ws.append(row)

		ws.cell(
			row=row_idx, column=column_idx_dict["Цена (по размерам)"]
		).alignment = wrap_alignment
		ws.cell(
			row=row_idx, column=column_idx_dict["Изображения"]
		).alignment = wrap_alignment

	for col in ws.columns:
		max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
		adjusted_width = min(max_len + 2, 60)

		ws.column_dimensions[col[0].column_letter].width = adjusted_width

	wb.save(filename)
	print(f"Report saved: {filename}")
